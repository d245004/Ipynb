import pandas as pd
from openpyxl import Workbook
from openpyxl.utils.dataframe import dataframe_to_rows
from openpyxl.styles import Font,Alignment,Border,Side,Color,PatternFill
import os
import warnings

# warning 문구 없애기
warnings.filterwarnings("ignore", message="Workbook contains no default style, apply openpyxl's default")
# pd.options.display.float_format = '{: .0f}'.format

pd.set_option('display.float_format', '{:,.0f}'.format)

imsi_test = pd.read_excel("C:/Users/Jaeri/Downloads/grid_exceldata.xlsx")

imsi_test['%'] = ''
imsi_test['접수번호'] = ''
imsi_test['공업사'] = ''
imsi_test['구분'] = ''
imsi_test['비고'] = ''

imsi_test['거래일시'] = imsi_test['거래일시'].str.slice(start=0, stop=10)
imsi_test['거래일시'] = imsi_test['거래일시'].str.replace(pat='.',repl='-',regex=False)
imsi_test = imsi_test.dropna(subset=['No'], how='any', axis=0)

wb = Workbook()

ws = wb.active
ws.title = 'Sheet1'

for r in dataframe_to_rows(imsi_test, index=False,header=True):
    ws.append(r)

# 테두리 box 변수
thin = Side(border_style="thin", color="000000")
double = Side(border_style="double", color="ff0000")
border_thin = Border(top=thin, left=thin, right=thin, bottom=thin)

m_row = ws.max_row

for row in range(1,m_row+1):
    ws.row_dimensions[row].height = 25

ws.column_dimensions['A'].width = 4
ws.column_dimensions['B'].width = 12
ws.column_dimensions['C'].width = 14
ws.column_dimensions['D'].width = 15
ws.column_dimensions['E'].width = 3
ws.column_dimensions['F'].width = 9
ws.column_dimensions['G'].width = 7
ws.column_dimensions['H'].width = 5
ws.column_dimensions['I'].width = 9

for row in ws['A1:I'+str(m_row)]:
    for cell in row:
        cell.border = border_thin

for row in ws['C2:C'+str(m_row)]:
    for cell in row:
        cell.number_format = '#,###'
        
for row in ws['A2:A'+str(m_row)]:
    for cell in row:
        cell.value = '=row()-1'
 
font_11 = Font(name='맑은 고딕',size=11,bold=True)

for row in ws[('C2:D'+str(m_row))]:
    for cell in row:
        cell.font = font_11

           
for row in ws[('A1:D'+str(m_row))]:
    for cell in row:
        cell.alignment = Alignment(horizontal='center',vertical='center')     

for row in ws[('C2:C'+str(m_row))]:
    for cell in row:
        cell.alignment = Alignment(horizontal='right',vertical='center')    


f_name = ws['B2'].value


ws.page_setup.paperSize = ws.PAPERSIZE_A4



ws.print_title_rows = '1:1'
ws.oddHeader.center.text = "보험건 입금 리스트"
ws.oddFooter.center.text =  "&[Page] / &N"

wb.save("D:/보험건 입금 리스트/"+f_name+" 보험건 입금 리스트.xlsx")

wb.close()

file_path = "C:/Users/Jaeri/Downloads/grid_exceldata.xlsx"
os.remove(file_path)



os.startfile("D:/보험건 입금 리스트/"+f_name+" 보험건 입금 리스트.xlsx")

# 매입 sheet 작업

# imsi_test